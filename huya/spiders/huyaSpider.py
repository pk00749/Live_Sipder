# -*- coding: utf-8 -*-
import sys
sys.path.append('..')
from scrapy.linkextractors import LinkExtractor
from scrapy.spiders import CrawlSpider, Rule
from scrapy.selector import Selector
from huya.items import HuyaItem
import urllib.request
import openpyxl
from bs4 import BeautifulSoup
from scrapy.http import Request
from huya.spiders.admin_excel import AdminWorkbook
from huya.spiders.get_rooms import GetRooms


class HuyaSpider(CrawlSpider):
    name = 'huya'
    total_rooms = 0
    no = 2
    file_path = 'G:\Program\Projects\Live_Sipder\huya.xlsx'
    get_room = GetRooms(file_path)
    workbook = AdminWorkbook(file_path)
    allowed_domains = ["huya.com"]

    start_urls = ['https://www.huya.com/g/']

    def start_requests(self):
        # url = self.get_room.get_topic_url(1+1)
        for url in self.start_urls:
            print('url: '+url)
            yield Request(url, callback=self.parse)


    def parse(self, response):
        self.logger.info("Visited %s", response.url)
        movies = response.xpath('//ul[@class="game-list clearfix"]/li')
        topic = self.workbook.read_cell('登录', 'C%d' % self.no)
        print('topic: ' + topic)
        self.workbook.create_sheet(topic)
        for each_movie in movies:
            if topic == each_movie.xpath('./a/img/@title').extract()[0]:
                href = each_movie.xpath('./a/@href').extract()[0]
                print('href: ' + href)
                yield Request(href, callback=self.get_topic_url)
            else:
                continue

    def get_topic_url(self, response):
        self.logger.info("Visited %s", response.url)
        live_rooms = response.xpath(
            '//ul[@class="live-list clearfix"]/li/a[@class="title new-clickstat"]/@href').extract()
        # total_rooms_in_page = len(live_rooms.find_all('a', class_='title new-clickstat', href=True))
        #
        # current_total_rooms_in_topic = self.total_rooms + total_rooms_in_page
        # print(live_rooms)
        for room in live_rooms:
            item = HuyaItem()
            item['room'] = room  # .xpath('.//a[@class="title new-clickstat"]/@herf')
            print(item['room'])














        # self.workbook.load_workbook()
        # movies = response.xpath('//ul[@class="game-list clearfix"]/li')
        # i = 1
        # for each_movie in movies:
        #     item = HuyaItem()
        #     item['topic'] = each_movie.xpath('./a/img/@title').extract()[0]
        #     item['href'] = each_movie.xpath('./a/@href').extract()[0]
        #     print(item['topic'])
        #     self.workbook.write_cell('主题列表', 'A%d' % i, item['topic'])
        #     self.workbook.write_cell('主题列表', 'B%d' % i, item['href'])
        #     i = i + 1
        #     yield item
        #     self.workbook.save_workbook()
