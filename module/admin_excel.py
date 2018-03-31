import openpyxl

class AdminWorkbook:

    def __init__(self, workbook_name):
        self.workbook_name = workbook_name
        self.workbook = self.load_workbook()

    def load_workbook(self):
        workbook = openpyxl.load_workbook(self.workbook_name)
        return workbook

    def read_cell(self, sheet_name, cell):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        return sheet[cell].value

    def write_cell(self, sheet_name, cell, value):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        sheet[cell] = value
        self.save_workbook()
#
    def save_workbook(self):
        self.workbook.save(self.workbook_name)

    def get_max_row(self, sheet_name):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        return sheet.max_row - 1
        # for row in sheet.iter_rows():
        #     for cell in row:
        #         print(cell.value)

if __name__ == '__main__':
    t = AdminWorkbook('../huya.xlsx')
    print(t.get_max_row('登录'))
    for i in range(1,2):
        print(i)
    # workbook = openpyxl.Workbook()
    # 写入文件
    # workbook.create_sheet('登录')
    # ws = workbook.create_sheet('房间')
    # test = admin_workbook('huya.xlsx')
    # test.load_workbook()
    # test.write_cell('房间', 'C2', 'www.baidu.com')
    # print(test.read_cell('登录', 'A1'))



# In [115]: sheetContent.get_highest_row()
#
# In [117]: sheetContent.get_highest_column()

# 新建文件
# workbook = openpyxl.Workbook()
 # 写入文件
# workbook.create_sheet('登录')
# workbook.create_sheet('房间')
# test = admin_workbook('huya.xlsx')
# test.load_workbook()
# test.write_cell('房间', 'C2', 'www.baidu.com')
# print(test.read_cell('登录', 'A1'))


# sheet_login = workbook.get_sheet_by_name('登录')
# print(sheet_login['A1'].value)
# workbook = openpyxl.load_workbook(os.path.join('../','huya.xlsx'))
# sheet_login = workbook.get_sheet_by_name('登录')
# sheet_login['A1'] = '登录名'
# sheet_login['B1'] = '密码'
#
#
#
# room = workbook.get_sheet_by_name('房间')
# # sheet.title = "New Shit"
# room['A1']='A1'
# room['C3'] = 'Hello world!'
# for i in range(10):
#     room["A%d" % (i+1)].value = i + 1
#
#
# print(sheet_login['A1'].value)
#  # 保存文件
# workbook.save('../huya.xlsx')


# import xlsxwriter as xw
# #新建excel
# workbook  = xw.Workbook('huya.xlsx')
# #新建工作薄
# sheet_login = workbook.add_worksheet()
# #写入数据
# sheet_login.write('A1', '登录名')
# sheet_login.write('B1', '密码')
# #关闭保存
# workbook.close()
